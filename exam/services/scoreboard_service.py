# services/scoreboard_service.py
import csv
import io
import os
from datetime import datetime
from django.core.files.base import ContentFile
from django.conf import settings
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from exam.models import UserExam,Exam
class ScoreboardService:
    """Service to handle scoreboard operations with CSV/Excel files"""
    
    @staticmethod
    def generate_scoreboard_csv(exam):
        """Generate CSV file for exam scoreboard"""
        # Get all completed exams
        user_exams = UserExam.objects.filter(
            exam=exam,
            is_completed=True
        ).select_related('profile__user').order_by('-percentage_score', 'completed_at')
        
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        writer.writerow([
            'Rank', 'Username', 'Name', 'Email', 'Score', 
            'Percentage Score', 'Completed At'
        ])
        
        # Write data
        current_rank = 1
        previous_score = None
        
        for idx, user_exam in enumerate(user_exams, 1):
            # Handle ties
            if previous_score != user_exam.percentage_score:
                current_rank = idx
            

            writer.writerow([
                current_rank,
                user_exam.profile.user.username,
                user_exam.profile.user_name or '',
                user_exam.profile.user.email,
                float(user_exam.score),
                float(user_exam.percentage_score),
                user_exam.completed_at.strftime('%Y-%m-%d %H:%M:%S') if user_exam.completed_at else ''
            ])
            
            previous_score = user_exam.percentage_score
        
        # Convert to ContentFile for saving
        csv_content = ContentFile(output.getvalue().encode('utf-8'))
        filename = f"scoreboard_{exam.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        # Save to exam's scoreboard_file field
        exam.scoreboard_file.save(filename, csv_content, save=True)
        
        return exam.scoreboard_file.path
    
    @staticmethod
    def generate_scoreboard_excel(exam):
        """Generate Excel file for exam scoreboard with formatting"""
        user_exams = UserExam.objects.filter(
            exam=exam,
            is_completed=True,
            score__isnull=False
        ).select_related('profile__user').order_by('-percentage_score', '-score', 'completed_at')
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Scoreboard - {exam.title[:31]}"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        headers = ['Rank', 'Username', 'Name', 'Email', 'Score', 'Percentage Score', 'Completed At']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data
        current_rank = 1
        previous_score = None
        row = 2
        
        for idx, user_exam in enumerate(user_exams, 1):
            if previous_score != user_exam.percentage_score:
                current_rank = idx

            ws.cell(row=row, column=1, value=current_rank)
            ws.cell(row=row, column=2, value=user_exam.profile.user.username)
            ws.cell(row=row, column=3, value=user_exam.profile.user_name or '')
            ws.cell(row=row, column=4, value=user_exam.profile.user.email)
            ws.cell(row=row, column=5, value=float(user_exam.score))
            ws.cell(row=row, column=6, value=float(user_exam.percentage_score))
            
            completed_at = user_exam.completed_at.strftime('%Y-%m-%d %H:%M:%S') if user_exam.completed_at else ''
            ws.cell(row=row, column=7, value=completed_at)
            
            # Color code based on percentage
            percent_cell = ws.cell(row=row, column=6)
            if user_exam.percentage_score >= 80:
                percent_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif user_exam.percentage_score >= 60:
                percent_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif user_exam.percentage_score < 60:
                percent_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            row += 1
            previous_score = user_exam.percentage_score
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to file
        filename = f"scoreboard_{exam.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(settings.MEDIA_ROOT, 'scoreboards', filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        wb.save(filepath)
        
        # Save reference in exam model
        exam.scoreboard_file.save(filename, open(filepath, 'rb'), save=True)
        
        return filepath
    
    @staticmethod
    def get_scoreboard_from_file(exam, format='csv'):
        """Read scoreboard from existing file or generate new one"""
        if exam.scoreboard_file and os.path.exists(exam.scoreboard_file.path):
            # Read existing file
            if format == 'csv':
                return ScoreboardService.read_csv_file(exam.scoreboard_file.path)
            else:
                return ScoreboardService.read_excel_file(exam.scoreboard_file.path)
        else:
            # Generate new file
            if format == 'csv':
                filepath = ScoreboardService.generate_scoreboard_csv(exam)
                return ScoreboardService.read_csv_file(filepath)
            else:
                filepath = ScoreboardService.generate_scoreboard_excel(exam)
                return ScoreboardService.read_excel_file(filepath)
    
    @staticmethod
    def read_csv_file(filepath):
        """Read CSV file and return data as list of dictionaries"""
        data = []
        with open(filepath, 'r', encoding='utf-8') as file:
            reader = csv.DictReader(file)
            for row in reader:
                data.append(row)
        return data
    
    @staticmethod
    def read_excel_file(filepath):
        """Read Excel file and return data as list of dictionaries"""
        df = pd.read_excel(filepath)
        return df.to_dict('records')
    
    @staticmethod
    def export_all_exams_scoreboard():
        """Export scoreboard for all exams"""
        exams = Exam.objects.all()
        results = {}
        
        for exam in exams:
            csv_path = ScoreboardService.generate_scoreboard_csv(exam)
            results[exam.id] = {
                'exam_title': exam.title,
                'file_path': csv_path,
                'participants': UserExam.objects.filter(exam=exam, is_completed=True).count()
            }
        
        return results